# add_columns.py — 向 Excel 追加/插入列, 表头样式继承参考列
#
# 用法:
#   1) add 模式(默认, add 可省略): 向 sheet 右侧追加列
#      python add_columns.py input.xlsx 完成日期:date 金额:currency "状态:select:通过,不通过,待确认" 数量:number:2 "小计:formula:=B{r}*C{r}"
#      python add_columns.py input.xlsx -s Sheet1 -s Sheet2 备注        # 只处理指定 sheet
#      python add_columns.py input.xlsx -o output.xlsx 完成日期:date     # 输出到新文件(默认原地修改)
#
#   2) left / right 模式: 在指定列的左侧/右侧插入列
#      python add_columns.py left  input.xlsx 金额 税前金额:currency    # 插入到 "金额" 列左侧
#      python add_columns.py right input.xlsx C 备注 备注2              # 插入到 C 列右侧
#      指定列可用: 表头名 / 列字母(如 C) / 列序号(从 1 开始)
#
#   3) index 模式: 向每个 sheet 最左列插入序号列, 数据行自动填补 1,2,3...
#      python add_columns.py index input.xlsx [列名]                    # 列名默认 "序号"
#
#   4) 修改下方 CONFIG 后直接运行(仅 add 模式; 命令行列定义优先于 CONFIG):
#      python add_columns.py input.xlsx
#
# 列定义格式: 名称[:类型[:选项或参数]]   (冒号中英文均可)
#   类型:
#     text        文本(默认), 只继承表头样式, 无附加格式
#     general     常规, 同 text
#     number      数值, 参数为小数位数(默认 2): 数量:number:2 -> 0.00; 数量:number:0 -> 整数
#     formula     公式, 参数为公式模板, {r} 代表当前行号: 小计:formula:=B{r}*C{r}
#     date        日期, 数字格式 yyyy-mm-dd
#     currency    货币(人民币), 数字格式 "¥"#,##0.00
#     select      单选, 第 2~1000 行加下拉框, 必须提供选项; 选项外用逗号输入会报错
#     multiselect 多选, 同样加下拉框, 但允许用英文逗号输入多个选项(提示但不拦截)
#   类型也接受中文: 文本 / 常规 / 数值 / 公式 / 日期 / 货币 / 单选 / 多选
#   选项用英文逗号分隔, 注意下拉公式总长不能超过 255 字符
#
# 样式继承规则:
#   - 只继承表头: 新列表头单元格复制参考列表头的全部样式(字体/底色/边框/对齐);
#     数据行不继承样式, 格式完全由列定义的类型决定(date/currency/number 覆盖数字格式)
#   - 参考列 = 插入位置左侧的原始列(add 模式为最右列; left/right 模式为锚定列或其左列,
#     在最左插入且无左列时退回锚定列本身)
#   - 连续添加多列时, 所有新列统一继承同一参考列, 不会级联继承前一个新增列
#   - 列宽按新列内容自适应: 表头名/下拉选项/类型格式样例的最宽者 + 2, 限定 10~30
#
# 边框选项(add/left/right 模式均可用, -b/--border):
#   - 不指定:  数据行无边框, 表头保留继承来的样式
#   - all:     新列所有单元格(表头 + 数据行)加细边框; 也可写 所有边框/所有
#   - none:    新列所有单元格去掉边框(含表头继承来的); 也可写 无边框/无
#
# 全列自适应(所有模式均可用, --autofit-all):
#   所有列按内容重算列宽, 类似 COM 的 Columns.AutoFit: 逐列扫描全部单元格,
#   按最宽内容 + 2 余量, 限定 10~60; 空列保持原宽. 默认只自适应新列.
#
# 依赖: pip install openpyxl
import argparse
import re
import sys
import unicodedata
from copy import copy
from pathlib import Path

# ============================================================
# CONFIG: 命令行未给列定义时使用(仅 add 模式); 命令行给了列定义则忽略此项
# ============================================================
CONFIG = {
    # 要追加的列, 格式同命令行列定义, 例如:
    # "columns": ["完成日期:date", "金额:currency", "状态:select:通过,不通过,待确认"],
    "columns": [],
    # 只处理这些 sheet; None 或空列表 = 全部 sheet
    "sheets": None,
    # 边框: None = 不处理; "all" = 所有边框; "none" = 无边框(命令行 -b 优先)
    "border": None,
    # 全列自适应: True = 所有列按内容重算列宽(命令行 --autofit-all 优先)
    "autofit_all": False,
}

TYPE_ALIASES = {
    "text": "text", "文本": "text",
    "general": "general", "常规": "general",
    "number": "number", "数值": "number",
    "formula": "formula", "公式": "formula",
    "date": "date", "日期": "date",
    "currency": "currency", "货币": "currency", "人民币": "currency",
    "select": "select", "单选": "select",
    "multiselect": "multiselect", "多选": "multiselect",
}
NUMBER_FORMATS = {
    "date": "yyyy-mm-dd",
    "currency": '"¥"#,##0.00',
}
DV_MAX_ROW = 1000          # 下拉框生效的行范围: 第 2 ~ 1000 行
BORDER_ALIASES = {
    "all": "all", "所有边框": "all", "所有": "all", "全部": "all",
    "none": "none", "无边框": "none", "无": "none",
}
TYPE_SAMPLE = {            # 类型的格式样例, 用于列宽自适应
    "date": "2026-01-01",
    "currency": "¥1,234,567.00",
}


def display_width(s):
    """中文按 2 个字符宽度算"""
    return sum(2 if unicodedata.east_asian_width(c) in "WF" else 1 for c in s)


def autofit_all_columns(ws, min_w=10, max_w=60):
    """类似 COM 的 Columns.AutoFit: 逐列扫描全部单元格内容, 按最宽内容重设列宽
    (+2 余量, 限定 min_w~max_w; 多行文本按最长行算; 空列保持原宽)"""
    from openpyxl.utils import get_column_letter

    for c in range(1, ws.max_column + 1):
        w = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            for line in str(v).splitlines():
                w = max(w, display_width(line))
        if w:
            ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, min_w), max_w)


def parse_column_spec(spec):
    """'名称:类型:选项或参数' -> (名称, 类型, [选项...], 参数)
    select/multiselect: 选项进 options; number: 参数=小数位数(int); formula: 参数=公式模板(str)"""
    parts = re.split(r"[:：]", spec)
    name = parts[0].strip()
    if not name:
        raise ValueError(f"列定义缺少名称: {spec!r}")
    raw_type = parts[1].strip() if len(parts) > 1 else "text"
    ctype = TYPE_ALIASES.get(raw_type.lower(), TYPE_ALIASES.get(raw_type))
    if ctype is None:
        raise ValueError(f"未知类型 {raw_type!r} (列 {name!r}), 支持: text/general/number/formula/date/currency/select/multiselect")
    tail = parts[2] if len(parts) > 2 else ""
    options, extra = [], None
    if ctype in ("select", "multiselect"):
        options = [o.strip() for o in re.split(r"[,，]", tail) if o.strip()]
        if not options:
            raise ValueError(f"{ctype} 类型必须提供选项, 如 '状态:{raw_type}:通过,不通过' (列 {name!r})")
        formula = ",".join(options)
        if len(formula) > 255:
            raise ValueError(f"列 {name!r} 的选项总长 {len(formula)} 超过 255 字符, 下拉框放不下")
    elif ctype == "number":
        extra = 2                                   # 默认 2 位小数
        if tail.strip():
            if not re.fullmatch(r"\d{1,2}", tail.strip()) or int(tail.strip()) > 10:
                raise ValueError(f"number 类型的小数位数须为 0~10, 如 '数量:number:2' (列 {name!r})")
            extra = int(tail.strip())
    elif ctype == "formula":
        extra = tail.strip()
        if not extra.startswith("="):
            raise ValueError(f"formula 类型必须以 = 开头, 如 '小计:formula:=B{{r}}*C{{r}}' (列 {name!r})")
    return name, ctype, options, extra


def resolve_column(ws, ref):
    """列引用 -> 1-based 列号; 支持表头名 / 列字母(A, B, AA) / 1-based 列序号"""
    from openpyxl.utils import column_index_from_string

    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip() == ref:
            return c
    if re.fullmatch(r"[A-Za-z]{1,3}", ref):
        idx = column_index_from_string(ref.upper())
    elif ref.isdigit():
        idx = int(ref)
    else:
        headers = [str(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
        raise ValueError(f"找不到列 {ref!r}, 现有表头: {headers} (也可用列字母或列序号)")
    if not 1 <= idx <= ws.max_column:
        raise ValueError(f"列 {ref!r} 超出范围, 该 sheet 共 {ws.max_column} 列")
    return idx


def parse_border(value):
    """边框选项 -> 'all' / 'none' / None"""
    if value is None:
        return None
    border = BORDER_ALIASES.get(str(value).strip().lower(), BORDER_ALIASES.get(str(value).strip()))
    if border is None:
        raise ValueError(f"未知边框选项 {value!r}, 支持: all/所有边框 或 none/无边框")
    return border


def apply_column(ws, col, spec, max_row, header_style, border=None):
    """写入一个新列: 表头样式+文字, 类型格式/公式, 下拉框, 边框, 列宽自适应"""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    name, ctype, options, extra = spec

    # 表头: 继承参考列表头样式 + 写表头文字
    if header_style is not None:
        ws.cell(row=1, column=col)._style = copy(header_style)
    ws.cell(row=1, column=col).value = name

    # date/currency: 数据区数字格式
    if ctype in NUMBER_FORMATS:
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col).number_format = NUMBER_FORMATS[ctype]

    # number: 限定小数位的数字格式
    if ctype == "number":
        fmt = "0" if extra == 0 else "0." + "0" * extra
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col).number_format = fmt

    # formula: 按模板填充公式, {r} 替换为当前行号
    if ctype == "formula":
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col).value = extra.replace("{r}", str(r))

    # 边框: all = 所有单元格加细边框; none = 去掉所有边框(含表头继承来的)
    if border is not None:
        from openpyxl.styles import Border, Side
        if border == "all":
            thin = Side(style="thin")
            b = Border(left=thin, right=thin, top=thin, bottom=thin)
        else:
            b = Border()
        for r in range(1, max_row + 1):
            ws.cell(row=r, column=col).border = b

    # 列宽自适应: 表头名 / 下拉选项 / 类型格式样例的最宽者 + 2, 限定 10~30
    letter = get_column_letter(col)
    candidates = [name, *options, TYPE_SAMPLE.get(ctype, "")]
    w = max(display_width(c) for c in candidates)
    ws.column_dimensions[letter].width = min(max(w + 2, 10), 30)

    # 单选/多选: 加下拉框
    if ctype in ("select", "multiselect"):
        dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                            allow_blank=True, showDropDown=False)
        # 注意: openpyxl 的 showDropDown=False 才是"显示下拉箭头"(Excel 里该属性含义相反)
        if ctype == "multiselect":
            dv.errorStyle = "information"      # 逗号输入多个时仅提示, 不拦截
            dv.errorTitle = "多选提示"
            dv.error = "多选列: 请从下拉框选择, 或用英文逗号分隔输入多个选项"
        else:
            dv.errorTitle = "输入无效"
            dv.error = "只能从下拉选项中选择一项"
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{DV_MAX_ROW}")

    return f"{name}({ctype})"


def load_targets(xlsx_path, sheets):
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    targets = sheets or wb.sheetnames
    for sname in targets:
        if sname not in wb.sheetnames:
            print(f"错误: 找不到 sheet {sname!r}, 现有 sheet: {wb.sheetnames}")
            sys.exit(1)
    return wb, targets


def save_workbook(wb, xlsx_path, output):
    out = Path(output) if output else xlsx_path
    wb.save(out)
    print(f"已保存 {out}" + ("" if out == xlsx_path else f" (源文件 {xlsx_path} 未改动)"))


def header_style_of(ws, col):
    """取某列表头单元格的样式副本, 无样式返回 None"""
    src = ws.cell(row=1, column=col)
    return copy(src._style) if src.has_style else None


DIM_PROPS = ("width", "hidden", "outlineLevel", "collapsed", "bestFit")


def capture_column_dimensions(ws):
    """抓取各列的维度属性(列宽/隐藏/分组等), 返回 {列号: {属性: 值}}
    展开 min~max 分组条目, 跳过全部属性为默认值的列; 列样式存为 _style 副本"""
    from openpyxl.utils import column_index_from_string

    saved = {}
    for letter, dim in list(ws.column_dimensions.items()):
        lo = dim.min or column_index_from_string(letter)
        hi = dim.max or lo
        vals = {p: getattr(dim, p, None) for p in DIM_PROPS}
        vals["_style"] = copy(dim._style) if getattr(dim, "_style", None) is not None else None
        if not any(v not in (None, False) for v in vals.values()):
            continue
        for c in range(lo, hi + 1):
            saved[c] = vals
    return saved


def restore_column_dimensions(ws, saved, insert_at, count):
    """insert_cols 不会平移 column_dimensions(openpyxl 3.1 实测), 需手动重建:
    insert_at 及之后的列定义右移 count 列, 之前的保持不变"""
    from openpyxl.utils import get_column_letter

    ws.column_dimensions.clear()
    for c, vals in saved.items():
        nc = c + count if c >= insert_at else c
        dim = ws.column_dimensions[get_column_letter(nc)]
        for p, v in vals.items():
            if p == "_style":
                if v is not None:
                    dim._style = copy(v)   # style 是只读别名, 直接写底层 _style
            else:
                setattr(dim, p, v)


def add_columns(xlsx_path, columns, sheets=None, output=None, border=None, autofit_all=False):
    """向 xlsx 各目标 sheet 右侧追加列, 表头样式统一继承原最右列"""
    xlsx_path = Path(xlsx_path)
    wb, targets = load_targets(xlsx_path, sheets)
    specs = [parse_column_spec(c) for c in columns]

    for sname in targets:
        ws = wb[sname]
        src_col = ws.max_column           # 参考列: 原最右列(固定, 不随新增列移动)
        style = header_style_of(ws, src_col)
        max_row = max(ws.max_row, 1)
        added = [apply_column(ws, src_col + 1 + i, spec, max_row, style, border)
                 for i, spec in enumerate(specs)]
        if autofit_all:
            autofit_all_columns(ws)
        print(f"  [{sname}] 追加 {len(added)} 列: {', '.join(added)}")

    save_workbook(wb, xlsx_path, output)


def insert_columns(xlsx_path, anchor, columns, side, sheets=None, output=None, border=None, autofit_all=False):
    """在各目标 sheet 的锚定列左侧/右侧插入列, 表头样式统一继承插入位置左侧的原列"""
    from openpyxl.utils import get_column_letter

    xlsx_path = Path(xlsx_path)
    wb, targets = load_targets(xlsx_path, sheets)
    specs = [parse_column_spec(c) for c in columns]

    for sname in targets:
        ws = wb[sname]
        try:
            anchor_idx = resolve_column(ws, anchor)
        except ValueError as e:
            print(f"错误: [{sname}] {e}")
            sys.exit(1)

        insert_at = anchor_idx if side == "left" else anchor_idx + 1
        # 参考列: 插入位置左侧的原列; 在最左插入且无左列时退回锚定列本身
        src_idx = insert_at - 1 if insert_at > 1 else anchor_idx
        style = header_style_of(ws, src_idx)
        max_row = max(ws.max_row, 1)

        # insert_cols 不平移 column_dimensions, 先抓取, 全部插入完成后重建右移
        saved_dims = capture_column_dimensions(ws)
        for i in range(len(specs)):
            ws.insert_cols(insert_at + i)
        restore_column_dimensions(ws, saved_dims, insert_at, len(specs))

        added = [apply_column(ws, insert_at + i, spec, max_row, style, border)
                 for i, spec in enumerate(specs)]
        if autofit_all:
            autofit_all_columns(ws)

        side_cn = "左侧" if side == "left" else "右侧"
        print(f"  [{sname}] 在 {anchor!r}({get_column_letter(anchor_idx)} 列){side_cn}插入 "
              f"{len(added)} 列: {', '.join(added)}")

    save_workbook(wb, xlsx_path, output)


def add_index_column(xlsx_path, name="序号", sheets=None, output=None, autofit_all=False):
    """在各目标 sheet 最左列插入序号列, 数据行自动填补 1,2,3..., 样式继承原 A 列"""
    xlsx_path = Path(xlsx_path)
    wb, targets = load_targets(xlsx_path, sheets)

    for sname in targets:
        ws = wb[sname]
        max_row = max(ws.max_row, 1)

        # 先取出原 A 列样式, 插入列后原 A 列右移为 B 列
        old_styles = [copy(ws.cell(row=r, column=1)._style) if ws.cell(row=r, column=1).has_style
                      else None for r in range(1, max_row + 1)]
        # insert_cols 不平移 column_dimensions, 先抓取, 插入后重建右移一列
        saved_dims = capture_column_dimensions(ws)
        ws.insert_cols(1)
        restore_column_dimensions(ws, saved_dims, 1, 1)
        for r in range(1, max_row + 1):
            if old_styles[r - 1] is not None:
                ws.cell(row=r, column=1)._style = old_styles[r - 1]

        ws.cell(row=1, column=1).value = name
        for r in range(2, max_row + 1):
            cell = ws.cell(row=r, column=1)
            cell.value = r - 1
            cell.number_format = "0"

        # 列宽自适应: 表头名 / 最大序号的较宽者 + 2, 限定 6~14
        w = max(display_width(name), len(str(max_row - 1)))
        ws.column_dimensions["A"].width = min(max(w + 2, 6), 14)
        if autofit_all:
            autofit_all_columns(ws)
        print(f"  [{sname}] 最左列插入序号列 {name!r}, 填充 1~{max_row - 1}")

    save_workbook(wb, xlsx_path, output)


COLUMN_EPILOG = ("列定义: 名称[:类型[:选项或参数]], 类型: text/general/number/formula/date/currency/select/multiselect "
                 "(或中文: 文本/常规/数值/公式/日期/货币/单选/多选)")


def add_border_arg(p):
    p.add_argument("-b", "--border", metavar="all|none",
                   help="边框: all/所有边框 = 新列所有单元格(含表头)加细边框; "
                        "none/无边框 = 去掉所有边框(含表头继承来的); 默认不处理")


def add_autofit_arg(p):
    p.add_argument("--autofit-all", action="store_true", dest="autofit_all",
                   help="所有列按内容重算列宽(类似 COM 的 Columns.AutoFit, +2 余量, 限定 10~60); "
                        "默认只自适应新列")


def border_or_exit(p, value):
    try:
        return parse_border(value)
    except ValueError as e:
        p.error(str(e))


def main(argv):
    args = list(argv)
    mode = "add"
    if args and args[0] in ("add", "left", "right", "index"):
        mode = args.pop(0)

    if mode == "index":
        p = argparse.ArgumentParser(
            prog="add_columns.py index",
            description="向每个 sheet 最左列插入序号列, 数据行自动填补 1,2,3...",
        )
        p.add_argument("xlsx", help="要修改的 xlsx 文件")
        p.add_argument("name", nargs="?", default="序号", help="序号列名, 默认 '序号'")
        p.add_argument("-s", "--sheet", action="append", dest="sheets",
                       help="只处理指定 sheet, 可多次使用; 默认全部 sheet")
        p.add_argument("-o", "--output", help="输出到新文件; 默认原地修改输入文件")
        add_autofit_arg(p)
        a = p.parse_args(args)
        add_index_column(a.xlsx, name=a.name, sheets=a.sheets, output=a.output,
                         autofit_all=a.autofit_all)
        return

    if mode in ("left", "right"):
        side_cn = "左" if mode == "left" else "右"
        p = argparse.ArgumentParser(
            prog=f"add_columns.py {mode}",
            description=f"在指定列{side_cn}侧插入列, 表头样式继承插入位置左侧的原列",
            epilog=COLUMN_EPILOG,
        )
        p.add_argument("xlsx", help="要修改的 xlsx 文件")
        p.add_argument("anchor", help="锚定列: 表头名 / 列字母(如 C) / 列序号(从 1 开始)")
        p.add_argument("columns", nargs="+",
                       help="列定义, 如 完成日期:date 数量:number:2 '状态:select:通过,不通过'")
        p.add_argument("-s", "--sheet", action="append", dest="sheets",
                       help="只处理指定 sheet, 可多次使用; 默认全部 sheet")
        p.add_argument("-o", "--output", help="输出到新文件; 默认原地修改输入文件")
        add_border_arg(p)
        add_autofit_arg(p)
        a = p.parse_args(args)
        insert_columns(a.xlsx, a.anchor, a.columns, side=mode, sheets=a.sheets, output=a.output,
                       border=border_or_exit(p, a.border), autofit_all=a.autofit_all)
        return

    p = argparse.ArgumentParser(
        prog="add_columns.py",
        description="向现有 Excel 右侧追加列, 表头样式继承原最右列; 在指定列左/右侧插入请用 left/right 模式",
        epilog=COLUMN_EPILOG,
    )
    p.add_argument("xlsx", help="要修改的 xlsx 文件")
    p.add_argument("columns", nargs="*",
                   help="列定义, 如 完成日期:date 数量:number:2 '小计:formula:=B{r}*C{r}' '状态:select:通过,不通过'")
    p.add_argument("-s", "--sheet", action="append", dest="sheets",
                   help="只处理指定 sheet, 可多次使用; 默认全部 sheet")
    p.add_argument("-o", "--output", help="输出到新文件; 默认原地修改输入文件")
    add_border_arg(p)
    add_autofit_arg(p)
    args = p.parse_args(args)

    columns = args.columns or CONFIG.get("columns") or []
    if not columns:
        p.print_help()
        print("\n错误: 没有给出任何列定义, 请在命令行提供或填写脚本内的 CONFIG['columns']")
        sys.exit(1)
    sheets = args.sheets or CONFIG.get("sheets") or None
    border = border_or_exit(p, args.border or CONFIG.get("border"))
    autofit_all = args.autofit_all or bool(CONFIG.get("autofit_all"))

    add_columns(args.xlsx, columns, sheets=sheets, output=args.output,
                border=border, autofit_all=autofit_all)


if __name__ == "__main__":
    main(sys.argv[1:])
