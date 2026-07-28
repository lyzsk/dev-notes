# md_to_excel.py — Markdown 转 Excel 工具, 支持两种模式
#
# 用法:
#   python md_to_excel.py [sheet] input.md [output.xlsx]       # sheet 模式(默认, 可省略 sheet)
#   python md_to_excel.py functionlist input.md [output.xlsx]  # 功能清单模式
#
# 依赖:
#   pip install pandas openpyxl      # sheet 模式
#   pip install openpyxl             # 功能清单模式仅需 openpyxl
#
# ============================================================
# 模式一: sheet — 通用 Markdown 表格转 Excel
#   1) 只有 `#` 标题(任意层级)和 `---` 分隔线才会开新 sheet
#   2) 空行隔开的连续表格属于同一个 sheet: 第一个是主表, 其后两列的表格视为
#      "列名 -> 说明" 对照表, 说明以批注形式挂到主表对应表头上
#   3) 列名 xxx(Y/N) -> 列名变为 xxx, 整列(第 2~1000 行)加 Y/N 下拉框
#   4) sheet 名取标题文本(无标题则 Table1/Table2...), 自动去非法字符/截断31字符/重名加后缀
#
# 模式二: functionlist — 功能清单范式 md 转 Excel (子项展开成行)
#   1) 每个 `## xxx` 生成一个 sheet; 标题以 "Function List" 结尾时 sheet 名取
#      前面的 xxx, 否则取整个 `##` 标题
#   2) 表头: 一级功能 / 二级功能 / 功能描述
#   3) 一级功能 = `#### N.M xxxx` 中的 xxxx (去掉编号); 若某个 `### N yyyy` 下
#      没有任何 `####`, 则回退用 yyyy 作为一级功能, 其下条目直接挂到该一级功能
#   4) 二级功能 = 顶格条目 `- **zzz**` 中的 zzz, 功能描述 = `: desc` 中的 desc
#   5) `- **zzz**:` 下的缩进子项 `- yyyy` 每个独立成行, 二级功能重复 zzz,
#      功能描述各取一条子项; 父条目本身无描述时不再单独占一行
#   6) 既无描述又无子项的条目, 功能描述留空
#   7) 样式: 内容区域全部边框; 首行底色 #153D63 + 白色加粗宋体 10 号居中;
#      其余行宋体 10 号, 自动换行, 列宽 30/36/100
# ============================================================
import re
import sys
import unicodedata
from pathlib import Path


def split_md_row(line):
    """按 | 切分一行，支持转义符 \\| 和行内代码 `...` 里的竖线"""
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    cells, buf, in_code, i = [], [], False, 0
    while i < len(line):
        ch = line[i]
        if ch == "\\" and i + 1 < len(line) and line[i + 1] == "|":
            buf.append("|")
            i += 2
            continue
        if ch == "`":
            in_code = not in_code
            i += 1
            continue
        if ch == "|" and not in_code:
            cells.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
        i += 1
    cells.append("".join(buf).strip())
    return cells


def is_separator(line):
    """| :-- | :-: | --: | 这种分隔行"""
    return bool(re.fullmatch(r"\|?[\s:\-|]+\|?", line.strip())) and "-" in line


YN_PATTERN = re.compile(r"\s*\(\s*[Yy]/[Nn]\s*\)\s*$")   # 列名结尾的 (Y/N)
HEADING_PATTERN = re.compile(r"^#{1,6}\s+(.*\S)\s*$")    # # ~ ###### 标题
HRULE_PATTERN = re.compile(r"^(-{3,}|\*{3,}|_{3,})$")    # --- / *** / ___ 分隔线


def build_table(rows):
    """把一个表格的行列表转成 (DataFrame, Y/N 列索引列表)"""
    import pandas as pd
    header, yn_cols = [], []
    for h in split_md_row(rows[0]):
        if YN_PATTERN.search(h):
            header.append(YN_PATTERN.sub("", h).strip())
            yn_cols.append(len(header) - 1)
        else:
            header.append(h)
    data = [split_md_row(r) for r in rows[1:] if not is_separator(r)]
    n = len(header)
    data = [r + [""] * (n - len(r)) if len(r) < n else r[:n] for r in data]
    return pd.DataFrame(data, columns=header), yn_cols


def parse_md_blocks(text):
    """按 # 标题 / --- 分隔线切 block。
    返回 [(heading, [(df, yn_cols), ...]), ...]
    同一 block 内空行隔开的多个表格同属一个 sheet。"""
    blocks, tables, rows, heading = [], [], [], None

    def flush_rows():
        nonlocal rows
        if len(rows) >= 2:
            tables.append(build_table(rows))
        rows = []

    def flush_block():
        nonlocal tables
        flush_rows()
        if tables:
            blocks.append((heading, tables))
        tables = []

    for line in text.splitlines():
        s = line.strip()
        m = HEADING_PATTERN.match(s)
        if m and not s.startswith("|"):
            flush_block()
            heading = m.group(1).strip()
        elif HRULE_PATTERN.match(s):
            flush_block()
            heading = None
        elif s.startswith("|") and s.endswith("|"):
            rows.append(s)
        else:
            flush_rows()          # 空行: 结束当前表格, 但不结束 block
    flush_block()
    return blocks


INVALID_SHEET_CHARS = re.compile(r'[\\/*?\[\]:]')        # Excel sheet 名非法字符


def make_sheet_name(title, used, idx):
    """sheet 名: 取标题文本, 去非法字符, 最长 31, 重名自动加 _2/_3 后缀"""
    name = INVALID_SHEET_CHARS.sub(" ", title).strip() if title else f"Table{idx}"
    name = (name or f"Table{idx}")[:31]
    base, k = name, 2
    while name in used:
        suffix = f"_{k}"
        name = base[:31 - len(suffix)] + suffix
        k += 1
    used.add(name)
    return name


def display_width(s):
    """中文按 2 个字符宽度算"""
    return sum(2 if unicodedata.east_asian_width(c) in "WF" else 1 for c in s)


def md_to_excel(md_path, xlsx_path=None, max_row=1000):
    """sheet 模式: 通用 Markdown 表格 -> Excel"""
    import pandas as pd
    md_path = Path(md_path)
    xlsx_path = xlsx_path or md_path.with_suffix(".xlsx")
    blocks = parse_md_blocks(md_path.read_text(encoding="utf-8"))
    if not blocks:
        print("未找到任何 Markdown 表格")
        return

    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    used_names = set()
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for i, (heading, tables) in enumerate(blocks, 1):
            df, yn_cols = tables[0]                       # 第一个表 = 主表
            sheet = make_sheet_name(heading, used_names, i)
            df.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets[sheet]

            for cell in ws[1]:                            # 表头加粗 + 底色 + 居中
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="2f5496")
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"                        # 冻结首行

            for col in ws.columns:                        # 列宽自适应
                w = max(display_width(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 8), 42)

            for c in yn_cols:                             # Y/N 列加下拉框
                letter = get_column_letter(c + 1)
                dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
                dv.error = "只能填写 Y 或 N"
                dv.errorTitle = "输入无效"
                dv.showErrorMessage = True
                ws.add_data_validation(dv)
                dv.add(f"{letter}2:{letter}{max_row}")

            # 后续表格 = 列名说明 -> 表头批注
            comments = 0
            for extra_df, _ in tables[1:]:
                if extra_df.shape[1] != 2:
                    print(f"  [{sheet}] 警告: 说明表不是两列, 已跳过")
                    continue
                desc_map = {str(r[0]).strip(): str(r[1]).strip()
                            for r in extra_df.itertuples(index=False) if str(r[0]).strip()}
                for c, col_name in enumerate(df.columns, 1):
                    if col_name in desc_map:
                        cm = Comment(desc_map[col_name], "列名说明")
                        cm.width, cm.height = 420, 160
                        ws.cell(row=1, column=c).comment = cm
                        comments += 1
            print(f"  [{sheet}] {df.shape[1]} 列, Y/N 下拉 {len(yn_cols)} 列, 表头批注 {comments} 个")
    print(f"已生成 {xlsx_path}，共 {len(blocks)} 个 sheet")


# ============================================================
# 模式二: functionlist — 功能清单范式 md 转 Excel
# ============================================================

FL_H2_RE = re.compile(r"^##\s+(.+?)\s*$")
FL_H3_RE = re.compile(r"^###\s+(.+?)\s*$")
FL_H4_RE = re.compile(r"^####\s+(.+?)\s*$")
FL_NUM_RE = re.compile(r"^\d+(\.\d+)*[\.、]?\s*")        # 标题前的编号, 如 "1." / "1.1"
FL_TOP_RE = re.compile(r"^- \*\*(.+?)\*\*\s*[:：]?\s*(.*)$")    # 顶格加粗条目
FL_SUB_RE = re.compile(r"^\s+-\s+(.*)$")                # 缩进子项


def parse_functionlist_block(block_lines):
    """解析一个 ####/### 块内的条目:
    '- **zzz**: desc' -> 一行 (zzz, desc)
    '- **zzz**:' + 缩进 '- yyyy' 子项 -> 每个子项一行, 二级功能重复 zzz
    既无描述又无子项的条目 -> 一行 (zzz, '')"""
    rows = []
    pending = None          # 最后一个无描述的加粗条目
    pending_has_subs = False
    for raw in block_lines:
        m = FL_TOP_RE.match(raw)
        if m:
            if pending and not pending_has_subs:
                rows.append((pending, ""))
            l2, desc = m.group(1).strip(), m.group(2).strip()
            if desc:
                rows.append((l2, desc))
                pending = None
            else:
                pending, pending_has_subs = l2, False
            continue
        m = FL_SUB_RE.match(raw)
        if m and pending:
            t = m.group(1).strip()
            if t:
                rows.append((pending, t))
                pending_has_subs = True
    if pending and not pending_has_subs:
        rows.append((pending, ""))
    return rows


def parse_functionlist_md(text):
    """返回 {sheet名: [(一级功能, 二级功能, 功能描述), ...]}
    每个 `## xxx` 一个 sheet; 一级功能取 `#### N.M`, 无 `####` 时回退取 `### N`"""
    sheets = {}
    sheet_name = None
    h3 = h4 = None
    h3_has_h4 = False
    h3_direct, h4_lines = [], []

    def strip_num(t):
        return FL_NUM_RE.sub("", t).strip()

    def flush_h4():
        nonlocal h4, h4_lines
        if h4 is not None:
            for l2, desc in parse_functionlist_block(h4_lines):
                sheets[sheet_name].append((strip_num(h4), l2, desc))
            h4, h4_lines = None, []

    def flush_h3():
        nonlocal h3, h3_has_h4, h3_direct
        flush_h4()
        if h3 is not None and not h3_has_h4:
            for l2, desc in parse_functionlist_block(h3_direct):
                sheets[sheet_name].append((strip_num(h3), l2, desc))
        h3, h3_has_h4, h3_direct = None, False, []

    for line in text.replace("\r\n", "\n").split("\n"):
        m = FL_H2_RE.match(line)
        if m:
            flush_h3()
            # sheet 名: 'EAP Function List' -> 'EAP'; 无该后缀则取整个标题
            sheet_name = re.sub(r"\s*Function List\s*$", "", m.group(1), flags=re.I).strip()
            sheet_name = sheet_name or m.group(1).strip()
            sheets.setdefault(sheet_name, [])
            continue
        if sheet_name is None:
            continue
        m = FL_H3_RE.match(line)
        if m:
            flush_h3()
            h3 = m.group(1).strip()
            continue
        m = FL_H4_RE.match(line)
        if m:
            flush_h4()
            h3_has_h4 = True
            h4 = m.group(1).strip()
            continue
        if h4 is not None:
            h4_lines.append(line)
        elif h3 is not None:
            h3_direct.append(line)
    flush_h3()
    return sheets


def functionlist_to_excel(md_path, xlsx_path=None):
    """功能清单模式: 功能清单范式 md -> Excel (子项展开成行, 支持 ### 回退)
    样式: 内容区域全部边框; 首行底色 #153d63 + 白色加粗宋体 10 号居中;
    其余行宋体 10 号自动换行; 列宽按内容自适应 (A 12~40, B 12~50, C 30~100)"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    md_path = Path(md_path)
    xlsx_path = xlsx_path or md_path.with_suffix(".xlsx")
    sheets = parse_functionlist_md(md_path.read_text(encoding="utf-8"))
    if not sheets:
        print("未找到任何 `## xxx` 章节")
        return

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="宋体", size=10, bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF153D63")
    header_align = Alignment(vertical="center", horizontal="center")
    body_font = Font(name="宋体", size=10, color="FF000000")
    body_align = Alignment(vertical="top", wrap_text=True)

    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        ws.append(["一级功能", "二级功能", "功能描述"])
        for row in rows:
            ws.append(row)
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
            for cell in r:
                cell.border = border
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                else:
                    cell.font = body_font
                    cell.alignment = body_align
        # 列宽按内容自适应: 中文按 2 字符宽, 多行取最长行; 各列设上下限, 超限靠换行
        col_bounds = ((12, 40), (12, 50), (30, 100))     # A 一级 / B 二级 / C 描述
        for (lo, hi), col_cells in zip(col_bounds, ws.iter_cols(max_col=3)):
            w = max(display_width(seg)
                    for c in col_cells for seg in str(c.value or "").split("\n"))
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(w + 2, lo), hi)
        print(f"  [{name}] {len(rows)} 行")

    wb.save(xlsx_path)
    print(f"已生成 {xlsx_path}，共 {len(sheets)} 个 sheet")


MODES = {
    "sheet": md_to_excel,
    "functionlist": functionlist_to_excel,
}


def main(argv):
    args = list(argv)
    mode = "sheet"
    if args and args[0] in MODES:
        mode = args.pop(0)
    if not args:
        print(__doc__ or "")
        print("用法: python md_to_excel.py [sheet|functionlist] input.md [output.xlsx]")
        sys.exit(1)
    MODES[mode](args[0], args[1] if len(args) > 1 else None)


if __name__ == "__main__":
    main(sys.argv[1:])
